# region [Imports]

import os
import sys
from gidtools.gidfiles import pathmaker, writejson, loadjson, writeit, pickleit
import datetime
import re
import statistics
import hashlib
from timeit import Timer
import configparser
from configparser import ExtendedInterpolation
from checksumdir import dirhash
import gidlogger as glog
from src.engine.config_reader import CfgSingletonProvider
from pprint import pformat
import csv
from openpyxl import Workbook
from copy import deepcopy
# endregion[Imports]

# region [Configs]

CFG = CfgSingletonProvider.get_config()

# endregion [Configs]

# region [Logging]

log = glog.aux_logger(__name__)
log.info(glog.imported(__name__))

# endregion[Logging]


def output_location(extension: str, suffix=None):
    _suffix = '' if suffix is None else f'_{suffix}'
    _name = CFG.get('general', 'output_file_name') + _suffix + '.' + extension
    _path = CFG.get_path('general', 'output_location').replace('cfg_dir', os.path.dirname(os.getenv('AS_INDEX_CFG_FILE')))
    if os.path.exists(_path) is False:
        if '.' not in os.path.basename(_path):
            os.makedirs(_path)
        else:
            raise AttributeError('path to create is not an folder but an file -- as it contains a "." in the name')

    return pathmaker(_path, _name)


def convert_extension_frequency_dict(extension_frequency_dict: dict, out_type: str = 'string'):
    # sourcery skip: list-comprehension
    if out_type == 'string':
        _temp_list = []
        for key, value in extension_frequency_dict.items():
            _temp_list.append(f"{key}: {str(value)}")
        return ' - '.join(_temp_list)


def printer_json_file(in_dict: dict):
    _output_file = output_location('json')
    writejson(in_dict, _output_file)


def printer_txt_file(in_dict: dict):
    _output_file = output_location('txt')
    writeit(_output_file, pformat(in_dict))


def printer_csv_file(in_dict: dict):
    in_dict = deepcopy(in_dict)
    _output_file_meta = output_location('csv', suffix='meta')
    _output_file_data = output_location('csv', suffix='data')
    if 'amount_of_files_per_extension' in in_dict.get('meta'):
        in_dict['meta']['amount_of_files_per_extension'] = convert_extension_frequency_dict(in_dict['meta']['amount_of_files_per_extension'], 'string')
    with open(_output_file_meta, 'w', newline='') as csvfile_meta:
        csv_writer = csv.writer(csvfile_meta, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)
        csv_writer.writerow([key for key in in_dict.get('meta')])
        csv_writer.writerow([value for key, value in in_dict.get('meta').items()])
    with open(_output_file_data, 'w', newline='') as csvfile_data:
        csv_writer = csv.writer(csvfile_data, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)
        csv_writer.writerow([key for key in in_dict['files'][0]])
        for item in in_dict.get('files'):
            csv_writer.writerow([value for key, value in item.items()])


def printer_pkl_file(in_dict: dict):
    _output_file = output_location('pkl')
    pickleit(in_dict, _output_file)


def printer_excel_file(in_dict: dict):
    in_dict = deepcopy(in_dict)
    _output_file = output_location('xlsx')
    if 'amount_of_files_per_extension' in in_dict.get('meta'):
        in_dict['meta']['amount_of_files_per_extension'] = convert_extension_frequency_dict(in_dict['meta']['amount_of_files_per_extension'], 'string')
    workb = Workbook()
    branch_sheet = workb.active
    branch_sheet.title = CFG.get('general', 'output_file_name')
    branch_sheet.cell(column=1, row=1, value='meta_key')
    branch_sheet.cell(column=2, row=1, value='meta_value')
    _row = 2
    for key, value in in_dict.get('meta').items():
        branch_sheet.cell(column=1, row=_row, value=key)
        branch_sheet.cell(column=2, row=_row, value=value)
        _row += 1
    _row += 4
    _temp_col = 1
    for key in in_dict.get('files')[0]:
        branch_sheet.cell(column=_temp_col, row=_row, value=key)
        _temp_col += 1
    _row += 1
    for item in in_dict.get('files'):
        _temp_col = 1
        for key, value in item.items():
            branch_sheet.cell(column=_temp_col, row=_row, value=value)
            _temp_col += 1
        _row += 1
    workb.save(_output_file)
